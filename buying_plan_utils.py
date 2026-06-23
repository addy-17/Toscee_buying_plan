"""
Buying Plan App - Shared Utilities
====================================
Scraping, inventory loading, Excel generation.
"""
import json, re, time, requests
from pathlib import Path
from datetime import datetime
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO

# --- File Paths ---
_APP_DIR = Path(__file__).parent
DATA_DIR = _APP_DIR.parent / "data"
CATALOG_FILE = DATA_DIR / "brand_catalogs.json"
CATALOG_FILE_LOCAL = _APP_DIR / "brand_catalogs.json"
if not CATALOG_FILE.exists() and CATALOG_FILE_LOCAL.exists():
    CATALOG_FILE = CATALOG_FILE_LOCAL
INVENTORY_JSON = _APP_DIR / "inventory_with_images.json"
BUYING_PLAN_EXCEL = _APP_DIR / "Toscee_buying_plan.xlsx"

# --- Category to Subcategory Mapping (kept for scraped data) ---
CATEGORY_MAP = {
    "Apparel": {"Sarees": r'\bsaree\b|\bsari\b', "Dresses": r'\bdress\b|\bgown\b|\bpartywear\b', "Kurta Sets": r'\bkurta\b|\bkurti\b|\bsharara\b', "Tops": r'\btop\b|\bshirt\b|\bblouse\b|\bcamisole\b', "Co-ord Sets": r'\bcoord\b|\bco-ord\b|\bcoordinate\b', "Kaftans": r'\bkaftan\b', "Ethnic Wear": r'\bethnic\b|\blehenga\b|\bdupatta\b'},
    "Jewellery": {"Earrings": r'\bearring\b|\bjhumka\b', "Necklaces": r'\bnecklace\b|\bneckpiece\b|\bchoker\b|\bpendant\b', "Bracelets & Bangles": r'\b(?:bracelet|bangle|hathphool)\b', "Rings": r'\bring\b', "Maang Tikka": r'\bmaang\s*tikka\b|\btikka\b'},
    "Fragrances": {"Perfumes": r'\bperfume\b|\battar\b|\bfragrance\b|\bscent\b', "Candles": r'\bcandle\b', "Incense": r'\bincense\b|\bagarbatti\b|\bdhoop\b'},
    "Bags": {"Bags": r'\bbag\b|\btote\b|\bclutch\b|\bpotli\b|\bhandbag\b'},
    "Home Decor": {"Bowls & Dips": r'\bbowl\b|\bdip\b', "Plates & Platters": r'\bplate\b|\bplatter\b|\bserveware\b', "Mugs": r'\bmug\b', "Coasters": r'\bcoaster\b', "Vases & Sculptures": r'\bvase\b|\bsculpture\b', "Table Runners": r'\brunner\b|\btable\s*mat\b'},
    "Accessories": {"Belts": r'\bbelt\b', "Scarves & Stoles": r'\bscarf\b|\bstole\b'},
    "Gifting": {"Gift Sets": r'\bgift\s*set\b|\bgift\s*box\b|\bhamper\b', "Stationery": r'\bjournal\b|\bstationery\b|\bnotebook\b|\bpen\b'},
}
ALL_SUBCATEGORIES = {}
for cat, subs in CATEGORY_MAP.items():
    ALL_SUBCATEGORIES.update(subs)
NAME_ONLY_SUBCATS = {"Sarees", "Dresses", "Kurta Sets", "Tops", "Co-ord Sets", "Kaftans", "Ethnic Wear", "Belts", "Scarves & Stoles"}

HEADERS_SCRAPE = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.5",
}

# --- Import the new loader ---
def load_inventory():
    """Load inventory from Toscee_buying_plan.xlsx by default."""
    try:
        from load_buying_plan_excel import load_buying_plan_excel as _load_new
        if BUYING_PLAN_EXCEL.exists():
            return _load_new(str(BUYING_PLAN_EXCEL))
    except Exception:
        pass
    
    # Fallback: load inventory_with_images.json
    inventory_path = Path(INVENTORY_JSON)
    if not inventory_path.exists():
        alt_paths = [
            _APP_DIR / "data" / "inventory_with_images.json",
            _APP_DIR.parent / "toscee_ai_mirror" / "data" / "inventory_with_images.json",
        ]
        for p in alt_paths:
            if p.exists():
                inventory_path = p
                break
    if inventory_path.exists():
        with open(inventory_path, "r", encoding="utf-8") as f:
            return json.load(f)
    return {"total_inventory": 0, "products": []}


def get_inventory_brands(inventory_data):
    """Return sorted list of unique brand names from inventory."""
    brands = set()
    for item in inventory_data.get("products", []):
        b = (item.get("brand") or "").strip()
        if b:
            brands.add(b)
    return sorted(brands)


def get_inventory_departments(inventory_data, brand_name):
    """Return sorted list of categories (Groups) for a given brand."""
    from load_buying_plan_excel import get_categories
    return get_categories(inventory_data, brand_name)


def get_inventory_products(inventory_data, brand_name, department=None):
    """Get products for a brand, optionally filtered by category/department."""
    products = []
    for item in inventory_data.get("products", []):
        b = (item.get("brand") or "").strip()
        if b.lower() == brand_name.lower():
            if department is None or (item.get("category") or "").lower() == department.lower():
                products.append(item)
    return products


def get_inventory_summary(inventory_data, brand_name):
    """Get summary info for a brand."""
    products = get_inventory_products(inventory_data, brand_name)
    depts = get_inventory_departments(inventory_data, brand_name)
    return {
        "brand_name": brand_name,
        "products_scraped": len(products),
        "departments": depts,
    }


# --- Legacy Catalog Loading (for scraped brands) ---

def load_catalog():
    if CATALOG_FILE.exists():
        with open(CATALOG_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {"brands": [], "total_brands_with_data": 0, "total_products_scraped": 0}


def save_catalog(catalog):
    catalog["scrape_date"] = datetime.now().strftime("%Y-%m-%d")
    catalog["scrape_timestamp"] = datetime.now().isoformat()
    with open(CATALOG_FILE, "w", encoding="utf-8") as f:
        json.dump(catalog, f, indent=2, ensure_ascii=False)


def get_brand_list(catalog):
    return sorted(set(b["brand_name"] for b in catalog.get("brands", [])))


# --- Product Matching (legacy) ---

def matches_subcategory(product, pattern):
    name = (product.get("product_name") or "").lower()
    return bool(re.search(pattern, name))


def matches_description(product, pattern):
    desc = (product.get("description") or "").lower()
    tags = product.get("tags", [])
    tags_str = " ".join(tags).lower() if isinstance(tags, list) else str(tags or "").lower()
    return bool(re.search(pattern, desc + " " + tags_str))


def product_matches_subcategory(product, subcat_name):
    pattern = ALL_SUBCATEGORIES.get(subcat_name)
    if not pattern:
        return False
    strict = subcat_name in NAME_ONLY_SUBCATS
    if matches_subcategory(product, pattern):
        return True
    if not strict and matches_description(product, pattern):
        return True
    return False


def get_subcategory_for_product(product):
    text = ((product.get("product_name") or "") + " " + (product.get("category") or "") + " " + (product.get("description") or "")).lower()
    tags = product.get("tags", [])
    text += " " + (" ".join(tags).lower() if isinstance(tags, list) else str(tags or "").lower())
    for subcat_name, pattern in ALL_SUBCATEGORIES.items():
        if re.search(pattern, text):
            return subcat_name
    return None


def get_category_for_subcategory(subcat_name):
    for cat_name, subs in CATEGORY_MAP.items():
        if subcat_name in subs:
            return cat_name
    return "Other"


# --- Scraping (kept for Add Brand tab) ---

def fetch_url(url, timeout=15):
    try:
        resp = requests.get(url, headers=HEADERS_SCRAPE, timeout=timeout)
        resp.raise_for_status()
        return resp
    except Exception:
        return None


def try_shopify_json(base_url):
    endpoints = ["/products.json?limit=250", "/collections/all/products.json?limit=250"]
    for endpoint in endpoints:
        url = base_url.rstrip("/") + endpoint
        resp = fetch_url(url)
        if resp and resp.status_code == 200:
            try:
                data = resp.json()
                products = data.get("products", [])
                if products:
                    return products
            except:
                continue
    return None


def parse_shopify_products(products, brand_url):
    result = []
    for p in products:
        title = p.get("title", "")
        variants = p.get("variants", [{}])
        price = variants[0].get("price") if variants else None
        images = p.get("images", [])
        img_url = images[0].get("src") if images else None
        tags_raw = p.get("tags", "")
        if isinstance(tags_raw, str):
            tags = [t.strip() for t in tags_raw.split(",") if t.strip()]
        elif isinstance(tags_raw, list):
            tags = tags_raw
        else:
            tags = []
        product_type = p.get("product_type", "")
        handle = p.get("handle", "")
        body_html = (p.get("body_html") or "")[:200]
        colours, sizes = [], []
        for option in p.get("options", []):
            opt_name = option.get("name", "").lower()
            if opt_name in ["color", "colour"]:
                colours = option.get("values", [])
            if opt_name in ["size", "sizes"]:
                sizes = option.get("values", [])
        product_url = f"{brand_url.rstrip('/')}/products/{handle}" if handle else ""
        result.append({
            "product_name": title, "price_mrp": float(price) if price else None,
            "compare_at_price": None, "category": product_type,
            "colours_available": colours, "sizes_available": sizes,
            "tags": tags, "description": BeautifulSoup(body_html, "html.parser").get_text()[:200] if body_html else "",
            "product_url": product_url, "image_url": img_url or "",
        })
    return result


def scrape_html_products(url):
    resp = fetch_url(url)
    if not resp:
        return []
    soup = BeautifulSoup(resp.text, "lxml")
    products = []
    product_selectors = [
        "div.product-item", "div.product-card", "li.product", "div.grid__item",
        "div.collection-product", "article.product", "div.product-loop__item",
        "div.product-grid-item", "div.item", "div[class*=product]",
        "li[class*=product]", "article[class*=product]",
        "div.thumbnail", "div.catalog-item"
    ]
    product_elements = []
    for selector in product_selectors:
        product_elements = soup.select(selector)
        if product_elements:
            break
    if not product_elements:
        product_links = soup.select('a[href*="/product"], a[href*="/products/"], a[href*="/collections/"]')
        seen = set()
        for link in product_links:
            href = link.get("href", "")
            if href and href not in seen and not href.startswith("#"):
                seen.add(href)
                name = link.get_text(strip=True) or ""
                if name and len(name) > 3:
                    products.append({
                        "product_name": name[:100], "price_mrp": None, "category": "",
                        "colours_available": [], "sizes_available": [], "description": "",
                        "product_url": href if href.startswith("http") else url.rstrip("/") + href,
                        "image_url": "",
                    })
        if products:
            return products
    for el in product_elements[:100]:
        name_el = el.select_one("h2, h3, h4, .product-title, .product-name, a[title]")
        name = name_el.get_text(strip=True) if name_el else ""
        price_el = el.select_one(".price, .product-price, .sale-price, span[class*=price]")
        price = None
        if price_el:
            price_text = price_el.get_text(strip=True)
            price_match = re.search(r'[\d,]+(?:\.\d{2})?', price_text.replace(",", ""))
            if price_match:
                price = float(price_match.group().replace(",", ""))
        link_el = (name_el.select_one("a") if name_el and name_el.name != "a" else (name_el or el.select_one("a")))
        product_url = ""
        if link_el and link_el.get("href"):
            href = link_el["href"]
            product_url = href if href.startswith("http") else url.rstrip("/") + "/" + href.lstrip("/")
        img_el = el.select_one("img")
        img_url = img_el.get("src") or img_el.get("data-src", "") if img_el else ""
        if name and len(name) > 3:
            products.append({
                "product_name": name[:100], "price_mrp": price, "category": "",
                "colours_available": [], "sizes_available": [], "description": "",
                "product_url": product_url, "image_url": img_url,
            })
    return products


def scrape_brand_by_url(url, brand_name=None):
    if not url or not url.startswith("http"):
        url = "https://" + url
    if not brand_name:
        from urllib.parse import urlparse
        parsed = urlparse(url)
        domain = parsed.netloc.replace("www.", "")
        brand_name = domain.split(".")[0].title()
    result = {"brand_name": brand_name, "website": url, "category": "Scraped", "products_scraped": 0, "scrape_method": "none", "scrape_status": "failed", "products": []}
    shopify_products = try_shopify_json(url)
    if shopify_products:
        products = parse_shopify_products(shopify_products, url)
        result["products"] = products; result["products_scraped"] = len(products); result["scrape_method"] = "shopify_api"; result["scrape_status"] = "success"
        return result
    html_products = scrape_html_products(url)
    if html_products:
        result["products"] = html_products; result["products_scraped"] = len(html_products); result["scrape_method"] = "html_scrape"; result["scrape_status"] = "success"
        return result
    for path in ["/collections/all", "/shop", "/collections", "/products", "/category"]:
        test_url = url.rstrip("/") + path
        html_products = scrape_html_products(test_url)
        if html_products:
            result["products"] = html_products; result["products_scraped"] = len(html_products); result["scrape_method"] = "html_scrape_deep"; result["scrape_status"] = "success"
            return result
    return result


# --- Excel Generation ---

def generate_buying_plan_excel(selected_items, total_budget):
    """
    Generate Excel from selected inventory items.
    selected_items: list of dicts from Toscee_buying_plan.xlsx, plus "quantity" key.
    
    Output columns matching the required format:
    Brand Name, Group, Sub Group, Product, Article, Product Title,
    Style Code /Sku, Color, Size, Barcode No, MRP, HSN Code, GST%,
    MATERIAL, Gender, Season, Quantity, PO RATE, UOM, MARGIN %, Net Price
    """
    wb = Workbook()
    
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2D3748", end_color="2D3748", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0")
    )
    data_font = Font(name="Calibri", size=10)
    summary_fill = PatternFill(start_color="48BB78", end_color="48BB78", fill_type="solid")
    
    # Columns matching the user's reference output
    headers_list = [
        "Brand Name", "Group", "Sub Group", "Product", "Article",
        "Product Title", "Style Code /Sku", "Color", "Size", "Barcode No",
        "MRP", "HSN Code", "GST%", "MATERIAL", "Gender", "Season",
        "Quantity", "PO RATE", "UOM", "MARGIN %", "Net Price"
    ]
    widths = [18, 22, 18, 18, 35, 50, 22, 14, 10, 18, 14, 16, 8, 14, 10, 10, 10, 14, 8, 10, 14]
    
    def write_headers(ws):
        for col, h in enumerate(headers_list, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
        for col, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = w
    
    def safe_float(val, default=None):
        if val is None:
            return default
        try:
            v = float(val)
            return v
        except:
            return default
    
    def get_field(item, field, default=""):
        val = item.get(field, default)
        if val is None:
            return default
        return val if val not in [None, "nan", "NaN"] else default

    def write_data_rows(ws, items):
        row = 2
        grand_qty = 0
        total_mrp = 0
        for item in items:
            qty = item.get("quantity", 0) or 0
            grand_qty += qty
            
            brand_name = get_field(item, "brand", "")
            category = get_field(item, "category", "")
            sub_category = get_field(item, "sub_category", "")
            product_type = get_field(item, "product", "")
            # Article = derived from category-sub_category-product
            article = f"{category}-{sub_category}-{product_type}" if sub_category else category
            
            product_title = get_field(item, "product_title", "")
            style_code = get_field(item, "style_code", "")
            color = get_field(item, "color", "")
            size = get_field(item, "size", "")
            barcode = get_field(item, "barcode", "")
            
            mrp = safe_float(get_field(item, "mrp"))
            hsn = get_field(item, "hsn_code", "")
            gst = get_field(item, "gst")
            material = get_field(item, "material", "")
            gender = get_field(item, "gender", "")
            season = get_field(item, "season", "")
            uom = get_field(item, "uom", "pcs")
            margin_pct = get_field(item, "margin_pct")
            net_price = safe_float(get_field(item, "net_price"))
            
            # PO RATE = MRP * 0.6 (legacy calculation, applied only if mrp exists)
            po_rate = round(mrp * 0.6, 2) if mrp is not None else None
            
            row_data = [
                brand_name, category, sub_category, product_type, article,
                product_title, style_code, color, size, barcode,
                round(mrp, 2) if mrp is not None else None,
                hsn, gst if gst is not None else None,
                material, gender, season,
                qty, round(po_rate, 2) if po_rate is not None else None,
                uom, margin_pct if margin_pct is not None else None,
                round(net_price, 2) if net_price is not None else None,
            ]
            
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.font = data_font
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center", wrap_text=(col == 6))
            
            if mrp is not None:
                total_mrp += mrp * qty
            row += 1
        return row, grand_qty
    
    def write_summary(ws, row, grand_qty, label="GRAND TOTAL"):
        row += 1
        ws.cell(row=row, column=6, value=label).font = Font(bold=True, size=11)
        total_cell = ws.cell(row=row, column=17, value=grand_qty)
        total_cell.font = Font(bold=True, size=11)
        total_cell.fill = summary_fill
        total_cell.alignment = Alignment(horizontal="center")
        return row
    
    # Group items by brand for per-brand sheets
    brand_groups = {}
    for item in selected_items:
        brand = get_field(item, "brand", "Unknown")
        if brand not in brand_groups:
            brand_groups[brand] = []
        brand_groups[brand].append(item)
    
    # Main sheet
    ws_main = wb.active
    ws_main.title = "Buying Plan"
    write_headers(ws_main)
    last_row, grand_qty = write_data_rows(ws_main, selected_items)
    write_summary(ws_main, last_row, grand_qty)
    ws_main.freeze_panes = "A2"
    ws_main.auto_filter.ref = f"A1:U{last_row + 1}"
    
    # Per-brand sheets
    for brand_name in sorted(brand_groups.keys()):
        ws = wb.create_sheet(title=brand_name[:31])
        write_headers(ws)
        last_row, brand_qty = write_data_rows(ws, brand_groups[brand_name])
        write_summary(ws, last_row, brand_qty, label=f"{brand_name} TOTAL")
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:U{last_row + 1}"
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"buying_plan_{timestamp}.xlsx"
    return output, filename